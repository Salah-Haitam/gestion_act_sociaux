"""Generation des exports Excel (openpyxl) et PDF (reportlab)."""

from __future__ import annotations

import io
from datetime import date
from decimal import Decimal

from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

ENTREPRISE = "MARSA MAROC"
SOUS_TITRE = "Direction des Ressources Humaines - Actions Sociales"

BLEU = "1F4E79"
BLEU_RL = colors.HexColor("#1F4E79")
GRIS_RL = colors.HexColor("#F2F5F9")


def _texte(valeur) -> str:
    if valeur is None:
        return ""
    if isinstance(valeur, Decimal):
        return f"{valeur:,.2f}".replace(",", " ")
    if isinstance(valeur, float):
        return f"{valeur:,.2f}".replace(",", " ")
    if isinstance(valeur, bool):
        return "Oui" if valeur else "Non"
    if isinstance(valeur, (list, tuple)):
        return ", ".join(str(v) for v in valeur)
    return str(valeur)


def export_excel(titre: str, colonnes: list[str], lignes: list[dict], nom_fichier: str) -> HttpResponse:
    """Construit un classeur Excel a partir d'une liste de dictionnaires."""
    wb = Workbook()
    ws = wb.active
    ws.title = titre[:31] or "Export"

    ws.append([f"{ENTREPRISE} - {titre}"])
    ws.append([f"{SOUS_TITRE} - edite le {date.today():%d/%m/%Y}"])
    ws.append([])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(len(colonnes), 1))
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=max(len(colonnes), 1))
    ws["A1"].font = Font(size=14, bold=True, color=BLEU)
    ws["A2"].font = Font(size=10, italic=True, color="666666")

    entete = ws.max_row + 1
    ws.append([c.replace("_", " ").title() for c in colonnes])
    fond = PatternFill("solid", fgColor=BLEU)
    for cellule in ws[entete]:
        cellule.font = Font(bold=True, color="FFFFFF")
        cellule.fill = fond
        cellule.alignment = Alignment(horizontal="center", vertical="center")

    for ligne in lignes:
        ws.append([_valeur_cellule(ligne.get(c)) for c in colonnes])

    for i, colonne in enumerate(colonnes, start=1):
        largeur = max(
            len(colonne) + 4,
            *(len(_texte(l.get(colonne))) + 3 for l in lignes[:200] or [{}]),
        )
        ws.column_dimensions[get_column_letter(i)].width = min(largeur, 45)
    ws.freeze_panes = ws.cell(row=entete + 1, column=1)

    flux = io.BytesIO()
    wb.save(flux)
    flux.seek(0)
    reponse = HttpResponse(
        flux.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    reponse["Content-Disposition"] = f'attachment; filename="{nom_fichier}.xlsx"'
    return reponse


def _valeur_cellule(valeur):
    """openpyxl ecrit nativement les nombres et les dates ; le reste en texte."""
    if isinstance(valeur, Decimal):
        return float(valeur)
    if isinstance(valeur, (int, float, date)) or valeur is None:
        return valeur
    if isinstance(valeur, (list, tuple)):
        return ", ".join(str(v) for v in valeur)
    return str(valeur)


def export_pdf(titre: str, colonnes: list[str], lignes: list[dict], nom_fichier: str) -> HttpResponse:
    """Construit un etat recapitulatif PDF en paysage."""
    flux = io.BytesIO()
    doc = SimpleDocTemplate(
        flux,
        pagesize=landscape(A4),
        leftMargin=1.2 * cm,
        rightMargin=1.2 * cm,
        topMargin=1.2 * cm,
        bottomMargin=1.2 * cm,
        title=titre,
    )
    styles = getSampleStyleSheet()
    style_titre = ParagraphStyle("titre", parent=styles["Title"], fontSize=16, textColor=BLEU_RL)
    style_sous = ParagraphStyle("sous", parent=styles["Normal"], fontSize=9, textColor=colors.grey)
    style_cell = ParagraphStyle("cell", parent=styles["Normal"], fontSize=7.5, leading=9)

    elements = [
        Paragraph(f"{ENTREPRISE} — {titre}", style_titre),
        Paragraph(f"{SOUS_TITRE} — edite le {date.today():%d/%m/%Y}", style_sous),
        Spacer(1, 0.5 * cm),
    ]

    if lignes:
        entetes = [
            Paragraph(f"<b>{c.replace('_', ' ').title()}</b>", style_cell) for c in colonnes
        ]
        donnees = [entetes] + [
            [Paragraph(_texte(l.get(c)), style_cell) for c in colonnes] for l in lignes
        ]
        largeur_dispo = doc.width
        table = Table(donnees, repeatRows=1, colWidths=[largeur_dispo / len(colonnes)] * len(colonnes))
        table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), BLEU_RL),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                    ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#B7C4D4")),
                    ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, GRIS_RL]),
                    ("LEFTPADDING", (0, 0), (-1, -1), 4),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 4),
                    ("TOPPADDING", (0, 0), (-1, -1), 3),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
                ]
            )
        )
        elements.append(table)
        elements.append(Spacer(1, 0.4 * cm))
        elements.append(Paragraph(f"Total : {len(lignes)} ligne(s).", style_sous))
    else:
        elements.append(Paragraph("Aucune donnee pour les criteres selectionnes.", styles["Normal"]))

    doc.build(elements)
    flux.seek(0)
    reponse = HttpResponse(flux.read(), content_type="application/pdf")
    reponse["Content-Disposition"] = f'attachment; filename="{nom_fichier}.pdf"'
    return reponse


def attestation_pdf(transaction) -> HttpResponse:
    """Attestation individuelle de versement d'une action sociale."""
    flux = io.BytesIO()
    doc = SimpleDocTemplate(
        flux,
        pagesize=A4,
        leftMargin=2.5 * cm,
        rightMargin=2.5 * cm,
        topMargin=2.5 * cm,
        bottomMargin=2.5 * cm,
        title="Attestation",
    )
    styles = getSampleStyleSheet()
    titre = ParagraphStyle("t", parent=styles["Title"], fontSize=15, textColor=BLEU_RL)
    corps = ParagraphStyle("c", parent=styles["Normal"], fontSize=11, leading=18)
    droite = ParagraphStyle("d", parent=corps, alignment=2)

    employe = transaction.matricule
    elements = [
        Paragraph(ENTREPRISE, titre),
        Paragraph(SOUS_TITRE, ParagraphStyle("s", parent=styles["Normal"], alignment=1, textColor=colors.grey)),
        Spacer(1, 1.2 * cm),
        Paragraph("<b>ATTESTATION DE PRESTATION SOCIALE</b>", ParagraphStyle("a", parent=titre, fontSize=13)),
        Spacer(1, 1 * cm),
        Paragraph(
            "Nous soussignes, Direction des Ressources Humaines de Marsa Maroc, attestons que :",
            corps,
        ),
        Spacer(1, 0.5 * cm),
    ]

    infos = [
        ["Matricule", employe.matricule],
        ["Nom et prenom", f"{employe.nom} {employe.prenom}"],
        ["Departement", employe.departement],
        ["Date de recrutement", f"{employe.date_recrutement:%d/%m/%Y}"],
        ["Action sociale", transaction.id_activitee.service],
        ["Montant verse", f"{transaction.montantTR:,.2f} MAD".replace(",", " ")],
        ["Date de versement", f"{transaction.date_transaction:%d/%m/%Y}"],
        ["Exercice", str(transaction.annee)],
        ["Reference", f"TR-{transaction.id_transaction:06d}"],
    ]
    table = Table(infos, colWidths=[5.5 * cm, 9 * cm])
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (0, -1), GRIS_RL),
                ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#B7C4D4")),
                ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 10),
                ("TOPPADDING", (0, 0), (-1, -1), 6),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
            ]
        )
    )
    elements += [
        table,
        Spacer(1, 1 * cm),
        Paragraph(
            "La presente attestation est delivree a l'interesse(e) pour servir et valoir ce que de droit.",
            corps,
        ),
        Spacer(1, 1.5 * cm),
        Paragraph(f"Fait a Casablanca, le {date.today():%d/%m/%Y}", droite),
        Spacer(1, 1.5 * cm),
        Paragraph("Le Directeur des Ressources Humaines", droite),
    ]

    doc.build(elements)
    flux.seek(0)
    reponse = HttpResponse(flux.read(), content_type="application/pdf")
    reponse["Content-Disposition"] = (
        f'attachment; filename="attestation_{employe.matricule}_{transaction.id_transaction}.pdf"'
    )
    return reponse
